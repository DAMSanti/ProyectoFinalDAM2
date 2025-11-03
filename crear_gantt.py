from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Diagrama de Gantt - ACEX"

# Colores del tema ACEX (azulados)
COLOR_PRINCIPAL = "1976D2"  # Azul principal
COLOR_OSCURO = "1565C0"     # Azul oscuro
COLOR_CLARO = "42A5F5"      # Azul claro
COLOR_MUY_CLARO = "BBDEF7"  # Azul muy claro (fondo)
COLOR_HEADER = "0D47A1"     # Azul muy oscuro para headers
COLOR_TEXTO_BLANCO = "FFFFFF"
COLOR_HITO = "FF9800"       # Naranja para hitos

# Estilos
header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
header_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=11)
fase_fill = PatternFill(start_color=COLOR_OSCURO, end_color=COLOR_OSCURO, fill_type="solid")
fase_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=10)
tarea_fill = PatternFill(start_color=COLOR_PRINCIPAL, end_color=COLOR_PRINCIPAL, fill_type="solid")
tarea_font = Font(color=COLOR_TEXTO_BLANCO, size=9)
subtarea_fill = PatternFill(start_color=COLOR_CLARO, end_color=COLOR_CLARO, fill_type="solid")
subtarea_font = Font(color=COLOR_TEXTO_BLANCO, size=9)
hito_fill = PatternFill(start_color=COLOR_HITO, end_color=COLOR_HITO, fill_type="solid")
hito_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=9)
barra_fill = PatternFill(start_color=COLOR_PRINCIPAL, end_color=COLOR_PRINCIPAL, fill_type="solid")
fondo_fill = PatternFill(start_color=COLOR_MUY_CLARO, end_color=COLOR_MUY_CLARO, fill_type="solid")

border_thin = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Headers principales
ws['A1'] = 'ID'
ws['B1'] = 'ACTIVIDAD'
ws['C1'] = 'INICIO'
ws['D1'] = 'FIN'
ws['E1'] = 'DURACIÓN'
ws['F1'] = 'RECURSOS'

# Aplicar estilo a headers
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws[f'{col}1'].fill = header_fill
    ws[f'{col}1'].font = header_font
    ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col}1'].border = border_thin

# Configurar anchos de columna
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 25

# Fechas del proyecto (16 semanas desde 01/09/2024)
fecha_inicio = datetime(2024, 9, 1)
semanas = 16

# Crear headers de semanas (columnas G en adelante)
col_offset = 7  # Columna G = 7
for semana in range(semanas):
    col_letter = get_column_letter(col_offset + semana)
    fecha_semana = fecha_inicio + timedelta(weeks=semana)
    ws[f'{col_letter}1'] = f'S{semana+1}'
    ws[f'{col_letter}2'] = fecha_semana.strftime('%d/%m')
    ws[f'{col_letter}1'].fill = header_fill
    ws[f'{col_letter}1'].font = header_font
    ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
    ws[f'{col_letter}1'].border = border_thin
    ws[f'{col_letter}2'].fill = PatternFill(start_color=COLOR_CLARO, end_color=COLOR_CLARO, fill_type="solid")
    ws[f'{col_letter}2'].font = Font(size=8, color=COLOR_TEXTO_BLANCO)
    ws[f'{col_letter}2'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col_letter}2'].border = border_thin
    ws.column_dimensions[col_letter].width = 3.5

# Datos del proyecto
actividades = [
    # ID, Nombre, Inicio (semana), Duración (semanas), Recursos, Tipo (fase/tarea/subtarea/hito)
    ["1", "■ ANÁLISIS Y DISEÑO", 0, 3, "Analista/Arquitecto", "fase"],
    ["1.1", "Análisis de requisitos", 0, 1, "Analista + Cliente", "tarea"],
    ["1.2", "Diseño de base de datos", 1, 1, "Arquitecto BD", "tarea"],
    ["1.3", "Diseño de arquitectura", 2, 1, "Arquitecto Software", "tarea"],
    ["H1", "🏁 HITO 1: Diseño Completado", 3, 0, "22/09/2024", "hito"],
    
    ["2", "■ DESARROLLO BACKEND", 2, 5, "Dev. Backend (2)", "fase"],
    ["2.1", "Configuración inicial", 2, 1, "Dev. Backend Senior", "tarea"],
    ["2.2", "Modelos y repositorios", 3, 1, "Dev. Backend (2)", "tarea"],
    ["2.3", "Servicios de negocio", 4, 2, "Dev. Backend (2)", "tarea"],
    ["2.4", "Controladores API", 6, 1, "Dev. Backend Senior", "tarea"],
    ["2.5", "Testing backend", 7, 1, "Dev. Backend + Tester", "tarea"],
    ["H2", "🏁 HITO 2: Backend Funcional", 8, 0, "27/10/2024", "hito"],
    
    ["3", "■ INTEGRACIÓN FIREBASE", 5, 3, "Dev. Backend + Cloud", "fase"],
    ["3.1", "Configuración Firebase", 5, 1, "Dev. Cloud", "tarea"],
    ["3.2", "Chat en tiempo real", 6, 1, "Dev. Backend", "tarea"],
    ["3.3", "Notificaciones push", 7, 1, "Dev. Backend", "tarea"],
    ["H3", "🏁 HITO 3: Firebase Integrado", 8, 0, "03/11/2024", "hito"],
    
    ["4", "■ DESARROLLO FRONTEND", 4, 6, "Dev. Frontend (2)", "fase"],
    ["4.1", "Configuración Flutter", 4, 1, "Dev. Frontend Senior", "tarea"],
    ["4.2", "Sistema de diseño", 5, 1, "Dev. Frontend + Designer", "tarea"],
    ["4.3", "Gestión de estado", 6, 1, "Dev. Frontend (2)", "tarea"],
    ["4.4", "Pantallas principales", 7, 2, "Dev. Frontend (2)", "tarea"],
    ["4.5", "Funcionalidades avanzadas", 9, 1, "Dev. Frontend (2)", "tarea"],
    ["4.6", "Testing frontend", 10, 1, "Dev. Frontend + Tester", "tarea"],
    ["H4", "🏁 HITO 4: App Completa", 11, 0, "17/11/2024", "hito"],
    
    ["5", "■ DESPLIEGUE E INFRAESTRUCTURA", 11, 2, "DevOps + Equipo", "fase"],
    ["5.1", "Configuración servidores", 11, 1, "DevOps", "tarea"],
    ["5.2", "Compilación aplicaciones", 12, 0.5, "Dev. Frontend", "tarea"],
    ["5.3", "Publicación", 12, 0.5, "DevOps + Project Mgr", "tarea"],
    ["H5", "🏁 HITO 5: Sistema en Producción", 13, 0, "01/12/2024", "hito"],
    
    ["6", "■ DOCUMENTACIÓN Y CIERRE", 11, 4, "Tech Writer + Equipo", "fase"],
    ["6.1", "Documentación técnica", 11, 1, "Tech Writer + Devs", "tarea"],
    ["6.2", "Documentación de usuario", 12, 1, "Tech Writer", "tarea"],
    ["6.3", "Entrega final", 14, 1, "Project Manager", "tarea"],
    ["H6", "🏁 HITO 6: Proyecto Finalizado", 15, 0, "22/12/2024", "hito"],
]

# Insertar actividades
fila = 3
for actividad in actividades:
    id_act, nombre, inicio_sem, duracion, recursos, tipo = actividad
    
    # Calcular fechas
    fecha_ini = fecha_inicio + timedelta(weeks=inicio_sem)
    if duracion > 0:
        fecha_fin = fecha_inicio + timedelta(weeks=inicio_sem + duracion)
        duracion_texto = f"{duracion} sem" if duracion >= 1 else f"{int(duracion*7)} días"
    else:
        fecha_fin = fecha_ini
        duracion_texto = "Hito"
    
    # Datos básicos
    ws[f'A{fila}'] = id_act
    ws[f'B{fila}'] = nombre
    ws[f'C{fila}'] = fecha_ini.strftime('%d/%m/%y')
    ws[f'D{fila}'] = fecha_fin.strftime('%d/%m/%y')
    ws[f'E{fila}'] = duracion_texto
    ws[f'F{fila}'] = recursos
    
    # Aplicar estilos según tipo
    if tipo == "fase":
        fill_style = fase_fill
        font_style = fase_font
    elif tipo == "tarea":
        fill_style = tarea_fill
        font_style = tarea_font
    elif tipo == "hito":
        fill_style = hito_fill
        font_style = hito_font
    else:
        fill_style = subtarea_fill
        font_style = subtarea_font
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{fila}'].fill = fondo_fill if tipo == "tarea" else fill_style
        ws[f'{col}{fila}'].font = Font(size=9) if tipo == "tarea" else font_style
        ws[f'{col}{fila}'].border = border_thin
        if col == 'B':
            ws[f'{col}{fila}'].alignment = Alignment(horizontal='left', vertical='center', indent=1 if tipo in ["tarea", "subtarea"] else 0)
        else:
            ws[f'{col}{fila}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Pintar barras de Gantt
    if duracion > 0:
        for sem in range(int(inicio_sem), int(inicio_sem + duracion * 4) + 1):  # *4 para cubrir mejor
            if sem < semanas:
                col_letra = get_column_letter(col_offset + sem)
                ws[f'{col_letra}{fila}'].fill = barra_fill
                ws[f'{col_letra}{fila}'].border = border_thin
    elif tipo == "hito":
        # Marcar hito con símbolo
        col_letra = get_column_letter(col_offset + int(inicio_sem))
        if int(inicio_sem) < semanas:
            ws[f'{col_letra}{fila}'] = "◆"
            ws[f'{col_letra}{fila}'].fill = hito_fill
            ws[f'{col_letra}{fila}'].font = Font(size=14, color=COLOR_TEXTO_BLANCO, bold=True)
            ws[f'{col_letra}{fila}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col_letra}{fila}'].border = border_thin
    
    fila += 1

# Añadir leyenda
fila_leyenda = fila + 2
ws[f'A{fila_leyenda}'] = "LEYENDA:"
ws[f'A{fila_leyenda}'].font = Font(bold=True, size=10)

ws[f'A{fila_leyenda+1}'] = "■"
ws[f'A{fila_leyenda+1}'].fill = fase_fill
ws[f'A{fila_leyenda+1}'].font = Font(size=14, color=COLOR_TEXTO_BLANCO)
ws[f'B{fila_leyenda+1}'] = "Fase principal"

ws[f'A{fila_leyenda+2}'] = "■"
ws[f'A{fila_leyenda+2}'].fill = tarea_fill
ws[f'A{fila_leyenda+2}'].font = Font(size=14, color=COLOR_TEXTO_BLANCO)
ws[f'B{fila_leyenda+2}'] = "Tarea"

ws[f'A{fila_leyenda+3}'] = "◆"
ws[f'A{fila_leyenda+3}'].fill = hito_fill
ws[f'A{fila_leyenda+3}'].font = Font(size=14, color=COLOR_TEXTO_BLANCO)
ws[f'B{fila_leyenda+3}'] = "Hito del proyecto"

# Añadir título del proyecto
ws.insert_rows(1)
ws.merge_cells('A1:F1')
ws['A1'] = "PROYECTO ACEX - DIAGRAMA DE GANTT"
ws['A1'].fill = PatternFill(start_color=COLOR_PRINCIPAL, end_color=COLOR_PRINCIPAL, fill_type="solid")
ws['A1'].font = Font(bold=True, size=14, color=COLOR_TEXTO_BLANCO)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

# Información del proyecto
ws.insert_rows(2)
ws.merge_cells('A2:F2')
ws['A2'] = "Duración: 16 semanas | Periodo: Septiembre - Diciembre 2024"
ws['A2'].fill = PatternFill(start_color=COLOR_CLARO, end_color=COLOR_CLARO, fill_type="solid")
ws['A2'].font = Font(size=10, color=COLOR_TEXTO_BLANCO)
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Congelar paneles
ws.freeze_panes = 'G4'

# Guardar archivo
archivo_salida = "g:/ProyectoFinalCSharp/ProyectoFinalDAM2/GANTT_ACEX.xlsx"
wb.save(archivo_salida)
print(f"✅ Diagrama de Gantt creado exitosamente: {archivo_salida}")
print(f"📊 Colores aplicados: Azules ACEX (#1976D2, #1565C0, #42A5F5)")
print(f"🎯 {len(actividades)} actividades incluidas con {len([a for a in actividades if a[5]=='hito'])} hitos")
