from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Recursos y Logística"

# Colores del tema ACEX (azulados)
COLOR_PRINCIPAL = "1976D2"      # Azul principal
COLOR_OSCURO = "1565C0"         # Azul oscuro
COLOR_CLARO = "42A5F5"          # Azul claro
COLOR_MUY_CLARO = "BBDEF7"      # Azul muy claro
COLOR_HEADER = "0D47A1"         # Azul muy oscuro
COLOR_SECCION = "1976D2"        # Azul para secciones
COLOR_SUBTOTAL = "64B5F6"       # Azul claro para subtotales
COLOR_TOTAL_FINAL = "0D47A1"    # Azul oscuro para total final
COLOR_TEXTO_BLANCO = "FFFFFF"
COLOR_FONDO_DATOS = "E3F2FD"    # Azul muy claro para datos

# Estilos
titulo_fill = PatternFill(start_color=COLOR_PRINCIPAL, end_color=COLOR_PRINCIPAL, fill_type="solid")
titulo_font = Font(bold=True, size=16, color=COLOR_TEXTO_BLANCO)

header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
header_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=11)

seccion_fill = PatternFill(start_color=COLOR_SECCION, end_color=COLOR_SECCION, fill_type="solid")
seccion_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=12)

subtotal_fill = PatternFill(start_color=COLOR_SUBTOTAL, end_color=COLOR_SUBTOTAL, fill_type="solid")
subtotal_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=11)

total_fill = PatternFill(start_color=COLOR_TOTAL_FINAL, end_color=COLOR_TOTAL_FINAL, fill_type="solid")
total_font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=14)

datos_fill = PatternFill(start_color=COLOR_FONDO_DATOS, end_color=COLOR_FONDO_DATOS, fill_type="solid")
datos_font = Font(size=10)

border_thin = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Función para aplicar estilo a una fila
def aplicar_estilo_fila(ws, fila, fill, font, cols):
    for col in cols:
        cell = ws[f'{col}{fila}']
        cell.fill = fill
        cell.font = font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center' if col != 'A' else 'left', vertical='center')

# Función para aplicar estilo a datos
def aplicar_estilo_datos(ws, fila, cols):
    for col in cols:
        cell = ws[f'{col}{fila}']
        cell.fill = datos_fill
        cell.font = datos_font
        cell.border = border_thin
        if col in ['E', 'F', 'G']:  # Columnas numéricas
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Configurar anchos de columna
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 18

# Título principal
fila = 1
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = 'RECURSOS Y LOGÍSTICA - PROYECTO ACEX'
ws[f'A{fila}'].fill = titulo_fill
ws[f'A{fila}'].font = titulo_font
ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[fila].height = 30

# Subtítulo
fila += 1
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = 'Resumen completo de recursos humanos, técnicos y logísticos'
ws[f'A{fila}'].fill = PatternFill(start_color=COLOR_CLARO, end_color=COLOR_CLARO, fill_type="solid")
ws[f'A{fila}'].font = Font(size=11, color=COLOR_TEXTO_BLANCO)
ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='center')

# Espacio
fila += 2

# =====================================================================
# SECCIÓN 1: RECURSOS HUMANOS
# =====================================================================
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = '📋 7.3.1 RECURSOS HUMANOS'
aplicar_estilo_fila(ws, fila, seccion_fill, seccion_font, ['A'])
ws.row_dimensions[fila].height = 25
fila += 1

# Headers
headers_rrhh = ['Rol', 'Cantidad', 'Dedicación', 'Periodo', 'Horas/Sem', 'Semanas', 'Coste Total']
for col_idx, header in enumerate(headers_rrhh, start=1):
    col_letter = get_column_letter(col_idx)
    ws[f'{col_letter}{fila}'] = header
aplicar_estilo_fila(ws, fila, header_fill, header_font, ['A', 'B', 'C', 'D', 'E', 'F', 'G'])
fila += 1

# Datos RRHH
datos_rrhh = [
    ['Project Manager', 1, 'Tiempo parcial (25%)', 'Todo el proyecto', 10, 16, '8.000 €'],
    ['Analista/Arquitecto', 1, 'Tiempo completo', 'Semanas 1-3', 40, 3, '6.600 €'],
    ['Dev. Backend Senior', 1, 'Tiempo completo', 'Semanas 3-11', 40, 8, '14.400 €'],
    ['Dev. Backend Junior', 1, 'Tiempo completo', 'Semanas 4-11', 40, 7, '8.400 €'],
    ['Dev. Frontend Senior', 1, 'Tiempo completo', 'Semanas 4-14', 40, 10, '18.000 €'],
    ['Dev. Frontend Junior', 1, 'Tiempo completo', 'Semanas 5-14', 40, 9, '10.800 €'],
    ['QA/Tester', 1, 'Tiempo parcial (50%)', 'Semanas 11-14', 20, 3, '2.100 €'],
    ['DevOps Engineer', 1, 'Tiempo parcial (50%)', 'Semanas 14-16', 20, 2, '2.000 €'],
    ['UI/UX Designer', 1, 'Tiempo parcial (25%)', 'Semanas 1-6', 10, 5, '2.000 €'],
    ['Technical Writer', 1, 'Tiempo parcial (50%)', 'Semanas 14-17', 20, 3, '2.100 €'],
]

for dato in datos_rrhh:
    for col_idx, valor in enumerate(dato, start=1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{fila}'] = valor
    aplicar_estilo_datos(ws, fila, ['A', 'B', 'C', 'D', 'E', 'F', 'G'])
    fila += 1

# Subtotal RRHH
ws.merge_cells(f'A{fila}:F{fila}')
ws[f'A{fila}'] = 'SUBTOTAL RECURSOS HUMANOS'
ws[f'G{fila}'] = '74.400 €'
aplicar_estilo_fila(ws, fila, subtotal_fill, subtotal_font, ['A', 'G'])
fila += 2

# =====================================================================
# SECCIÓN 2: HARDWARE
# =====================================================================
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = '💻 7.3.2 RECURSOS TÉCNICOS (HARDWARE)'
aplicar_estilo_fila(ws, fila, seccion_fill, seccion_font, ['A'])
ws.row_dimensions[fila].height = 25
fila += 1

# Headers
headers_hw = ['Recurso', 'Cantidad', 'Uso', '', 'Coste Unit.', '', 'Coste Total']
for col_idx, header in enumerate(headers_hw, start=1):
    if header:
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{fila}'] = header
ws.merge_cells(f'C{fila}:D{fila}')
ws.merge_cells(f'E{fila}:F{fila}')
aplicar_estilo_fila(ws, fila, header_fill, header_font, ['A', 'B', 'C', 'E', 'G'])
fila += 1

# Datos Hardware
datos_hw = [
    ['Portátil Dev (Windows)', 4, 'Desarrollo backend/frontend', '', '1.200 €', '', '4.800 €'],
    ['Portátil Dev (MacBook Pro)', 2, 'Desarrollo iOS', '', '2.500 €', '', '5.000 €'],
    ['Servidor local desarrollo', 1, 'Testing y pruebas', '', '1.500 €', '', '1.500 €'],
    ['iPhone (testing iOS)', 1, 'Testing aplicación iOS', '', '800 €', '', '800 €'],
    ['Android devices (varios)', 3, 'Testing aplicación Android', '', '300 €', '', '900 €'],
    ['Tablet Android', 1, 'Testing UI responsive', '', '400 €', '', '400 €'],
    ['Monitor adicional', 6, 'Mejora productividad', '', '200 €', '', '1.200 €'],
    ['Almacenamiento NAS', 1, 'Backup y compartir archivos', '', '600 €', '', '600 €'],
]

for dato in datos_hw:
    ws[f'A{fila}'] = dato[0]
    ws[f'B{fila}'] = dato[1]
    ws.merge_cells(f'C{fila}:D{fila}')
    ws[f'C{fila}'] = dato[2]
    ws.merge_cells(f'E{fila}:F{fila}')
    ws[f'E{fila}'] = dato[4]
    ws[f'G{fila}'] = dato[6]
    aplicar_estilo_datos(ws, fila, ['A', 'B', 'C', 'E', 'G'])
    fila += 1

# Subtotal Hardware
ws.merge_cells(f'A{fila}:F{fila}')
ws[f'A{fila}'] = 'SUBTOTAL HARDWARE'
ws[f'G{fila}'] = '15.200 €'
aplicar_estilo_fila(ws, fila, subtotal_fill, subtotal_font, ['A', 'G'])
fila += 2

# =====================================================================
# SECCIÓN 3: SOFTWARE Y SERVICIOS
# =====================================================================
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = '🔧 7.3.3 RECURSOS TÉCNICOS (SOFTWARE Y SERVICIOS)'
aplicar_estilo_fila(ws, fila, seccion_fill, seccion_font, ['A'])
ws.row_dimensions[fila].height = 25
fila += 1

# Headers
headers_sw = ['Recurso', 'Tipo', 'Uso', 'Coste Mensual', '', '', 'Coste Total (4m)']
for col_idx, header in enumerate(headers_sw, start=1):
    if header:
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{fila}'] = header
ws.merge_cells(f'C{fila}:D{fila}')
ws.merge_cells(f'E{fila}:F{fila}')
aplicar_estilo_fila(ws, fila, header_fill, header_font, ['A', 'B', 'C', 'E', 'G'])
fila += 1

# Datos Software
datos_sw = [
    ['Visual Studio Professional', 'Licencia', 'IDE backend', '45 € × 2 dev', '', '', '360 €'],
    ['JetBrains IntelliJ/Rider', 'Licencia', 'IDE alternativo', '24 € × 2 dev', '', '', '192 €'],
    ['GitHub Pro', 'Suscripción', 'Control de versiones', '4 € × 10 users', '', '', '160 €'],
    ['Azure SQL Database', 'Cloud', 'Base de datos desarrollo', '50 €', '', '', '200 €'],
    ['Azure App Service', 'Cloud', 'Hosting API desarrollo', '40 €', '', '', '160 €'],
    ['Firebase Blaze Plan', 'Cloud', 'Firestore + Storage + FCM', '30 €', '', '', '120 €'],
    ['Google Play Console', 'Pago único', 'Publicación Android', '-', '', '', '25 €'],
    ['Apple Developer Program', 'Anual', 'Publicación iOS', '99 €', '', '', '99 €'],
    ['Figma Pro', 'Suscripción', 'Diseño UI/UX', '12 €', '', '', '48 €'],
    ['Postman Team', 'Suscripción', 'Testing APIs', '24 €', '', '', '96 €'],
    ['Jira Software', 'Suscripción', 'Gestión de proyecto', '10 € × 10 users', '', '', '400 €'],
    ['Slack Pro', 'Suscripción', 'Comunicación equipo', '6 € × 10 users', '', '', '240 €'],
    ['Office 365 Business', 'Suscripción', 'Documentación', '10 € × 10 users', '', '', '400 €'],
]

for dato in datos_sw:
    ws[f'A{fila}'] = dato[0]
    ws[f'B{fila}'] = dato[1]
    ws.merge_cells(f'C{fila}:D{fila}')
    ws[f'C{fila}'] = dato[2]
    ws.merge_cells(f'E{fila}:F{fila}')
    ws[f'E{fila}'] = dato[3]
    ws[f'G{fila}'] = dato[6]
    aplicar_estilo_datos(ws, fila, ['A', 'B', 'C', 'E', 'G'])
    fila += 1

# Subtotal Software
ws.merge_cells(f'A{fila}:F{fila}')
ws[f'A{fila}'] = 'SUBTOTAL SOFTWARE Y SERVICIOS (4 meses)'
ws[f'G{fila}'] = '2.500 €'
aplicar_estilo_fila(ws, fila, subtotal_fill, subtotal_font, ['A', 'G'])
fila += 2

# =====================================================================
# SECCIÓN 4: INFRAESTRUCTURA DE PRODUCCIÓN
# =====================================================================
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = '☁️ 7.3.4 INFRAESTRUCTURA DE PRODUCCIÓN'
aplicar_estilo_fila(ws, fila, seccion_fill, seccion_font, ['A'])
ws.row_dimensions[fila].height = 25
fila += 1

# Headers
headers_infra = ['Recurso', 'Proveedor', 'Especificaciones', 'Coste Mensual', '', '', 'Coste Anual']
for col_idx, header in enumerate(headers_infra, start=1):
    if header:
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{fila}'] = header
ws.merge_cells(f'C{fila}:D{fila}')
ws.merge_cells(f'E{fila}:F{fila}')
aplicar_estilo_fila(ws, fila, header_fill, header_font, ['A', 'B', 'C', 'E', 'G'])
fila += 1

# Datos Infraestructura
datos_infra = [
    ['SQL Server Database', 'Azure', 'Standard S2 (50 DTUs)', '75 €', '', '', '900 €'],
    ['App Service (API)', 'Azure', 'Premium P1V2', '140 €', '', '', '1.680 €'],
    ['Firebase Hosting', 'Google', 'Blaze Plan (uso moderado)', '50 €', '', '', '600 €'],
    ['CDN (imágenes)', 'Cloudflare', 'Pro Plan', '20 €', '', '', '240 €'],
    ['Dominio .com', 'GoDaddy', 'Registro anual', '-', '', '', '12 €'],
    ['SSL Certificate', 'Let\'s Encrypt', 'Gratuito', '0 €', '', '', '0 €'],
    ['Backup Storage', 'Azure Blob', '100 GB redundante', '5 €', '', '', '60 €'],
    ['Monitoring (App Insights)', 'Azure', 'Uso básico', '15 €', '', '', '180 €'],
]

for dato in datos_infra:
    ws[f'A{fila}'] = dato[0]
    ws[f'B{fila}'] = dato[1]
    ws.merge_cells(f'C{fila}:D{fila}')
    ws[f'C{fila}'] = dato[2]
    ws.merge_cells(f'E{fila}:F{fila}')
    ws[f'E{fila}'] = dato[3]
    ws[f'G{fila}'] = dato[6]
    aplicar_estilo_datos(ws, fila, ['A', 'B', 'C', 'E', 'G'])
    fila += 1

# Subtotal Infraestructura
ws.merge_cells(f'A{fila}:F{fila}')
ws[f'A{fila}'] = 'SUBTOTAL INFRAESTRUCTURA (primer año)'
ws[f'G{fila}'] = '3.672 €'
aplicar_estilo_fila(ws, fila, subtotal_fill, subtotal_font, ['A', 'G'])
fila += 2

# =====================================================================
# SECCIÓN 5: ESPACIOS Y LOGÍSTICA
# =====================================================================
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = '🏢 7.3.5 ESPACIOS Y LOGÍSTICA'
aplicar_estilo_fila(ws, fila, seccion_fill, seccion_font, ['A'])
ws.row_dimensions[fila].height = 25
fila += 1

# Headers
headers_log = ['Recurso', 'Tipo', 'Cantidad', 'Coste Mensual', '', '', 'Coste Total (4m)']
for col_idx, header in enumerate(headers_log, start=1):
    if header:
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{fila}'] = header
ws.merge_cells(f'C{fila}:D{fila}')
ws.merge_cells(f'E{fila}:F{fila}')
aplicar_estilo_fila(ws, fila, header_fill, header_font, ['A', 'B', 'C', 'E', 'G'])
fila += 1

# Datos Logística
datos_log = [
    ['Espacio de oficina', 'Alquiler', '50 m²', '800 €', '', '', '3.200 €'],
    ['Internet de alta velocidad', 'Servicio', '1 línea 600 Mbps', '60 €', '', '', '240 €'],
    ['Electricidad y servicios', 'Servicios', '-', '150 €', '', '', '600 €'],
    ['Mobiliario (mesas, sillas)', 'Compra', 'Para 10 personas', '-', '', '', '2.500 €'],
    ['Material de oficina', 'Consumibles', '-', '50 €', '', '', '200 €'],
    ['Café y snacks', 'Beneficios', '-', '100 €', '', '', '400 €'],
]

for dato in datos_log:
    ws[f'A{fila}'] = dato[0]
    ws[f'B{fila}'] = dato[1]
    ws.merge_cells(f'C{fila}:D{fila}')
    ws[f'C{fila}'] = dato[2]
    ws.merge_cells(f'E{fila}:F{fila}')
    ws[f'E{fila}'] = dato[3]
    ws[f'G{fila}'] = dato[6]
    aplicar_estilo_datos(ws, fila, ['A', 'B', 'C', 'E', 'G'])
    fila += 1

# Subtotal Logística
ws.merge_cells(f'A{fila}:F{fila}')
ws[f'A{fila}'] = 'SUBTOTAL ESPACIOS Y LOGÍSTICA (4 meses)'
ws[f'G{fila}'] = '7.140 €'
aplicar_estilo_fila(ws, fila, subtotal_fill, subtotal_font, ['A', 'G'])
fila += 2

# =====================================================================
# TOTAL GENERAL
# =====================================================================
ws.merge_cells(f'A{fila}:F{fila}')
ws[f'A{fila}'] = '💰 TOTAL RECURSOS Y LOGÍSTICA'
ws[f'G{fila}'] = '102.912 €'
aplicar_estilo_fila(ws, fila, total_fill, total_font, ['A', 'G'])
ws.row_dimensions[fila].height = 30

# Resumen en la parte inferior
fila += 3
ws.merge_cells(f'A{fila}:G{fila}')
ws[f'A{fila}'] = '📊 RESUMEN POR CATEGORÍAS'
aplicar_estilo_fila(ws, fila, PatternFill(start_color=COLOR_CLARO, end_color=COLOR_CLARO, fill_type="solid"), 
                     Font(bold=True, size=11, color=COLOR_TEXTO_BLANCO), ['A'])
fila += 1

# Tabla resumen
resumen = [
    ['Recursos Humanos', '74.400 €', '72.3%'],
    ['Hardware', '15.200 €', '14.8%'],
    ['Software y Servicios (4 meses)', '2.500 €', '2.4%'],
    ['Infraestructura (primer año)', '3.672 €', '3.6%'],
    ['Espacios y Logística (4 meses)', '7.140 €', '6.9%'],
]

ws[f'A{fila}'] = 'Categoría'
ws[f'B{fila}'] = 'Importe'
ws[f'C{fila}'] = '% del Total'
ws.merge_cells(f'C{fila}:G{fila}')
aplicar_estilo_fila(ws, fila, header_fill, header_font, ['A', 'B', 'C'])
fila += 1

for item in resumen:
    ws[f'A{fila}'] = item[0]
    ws[f'B{fila}'] = item[1]
    ws.merge_cells(f'C{fila}:G{fila}')
    ws[f'C{fila}'] = item[2]
    aplicar_estilo_datos(ws, fila, ['A', 'B', 'C'])
    fila += 1

# Congelar paneles
ws.freeze_panes = 'A6'

# Guardar archivo
archivo_salida = "g:/ProyectoFinalCSharp/ProyectoFinalDAM2/RECURSOS_LOGISTICA_ACEX.xlsx"
wb.save(archivo_salida)
print(f"✅ Recursos y Logística creado exitosamente: {archivo_salida}")
print(f"📊 Secciones incluidas:")
print(f"   • Recursos Humanos: 74.400 €")
print(f"   • Hardware: 15.200 €")
print(f"   • Software y Servicios: 2.500 €")
print(f"   • Infraestructura: 3.672 €")
print(f"   • Espacios y Logística: 7.140 €")
print(f"💰 TOTAL: 102.912 €")
